"""
Gerenciamento de Rotina Logística BURN RJ
Painel de Recebimento de Produto Acabado
Backend Flask — serve o dashboard e as APIs de dados.
"""
import logging
import os
import json
import io
import base64
import threading
import time
from datetime import datetime, timezone, timedelta

import requests as _requests
import qrcode
from flask import Flask, jsonify, request, send_from_directory, Response
from flask_cors import CORS

from receipt_collector import ReceiptCollector, DEPOSITOS, STATUS_ORDER

_BRT = timezone(timedelta(hours=-3))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
POLL_INTERVALO = 30  # segundos

app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="/static")
CORS(app)

collector = ReceiptCollector(intervalo=POLL_INTERVALO)


def _self_ping():
    """Faz GET no próprio /api/status a cada 10 min para evitar hibernação do Render."""
    # Aguarda 2 min para o app subir completamente antes do primeiro ping
    time.sleep(120)
    host = os.environ.get("RENDER_EXTERNAL_URL", "http://localhost:5002")
    url  = f"{host}/api/status"
    while True:
        try:
            _requests.get(url, timeout=10)
            log.debug(f"Self-ping OK: {url}")
        except Exception as e:
            log.debug(f"Self-ping falhou (normal na inicializacao): {e}")
        time.sleep(240)  # 4 minutos — evita hibernação no Render free tier


_ping_thread = threading.Thread(target=_self_ping, daemon=True, name="self-ping")
_ping_thread.start()


@app.before_request
def _garantir_collector():
    collector.iniciar()  # idempotente — inicia thread no worker certo após gunicorn fork


# ─────────────────────────────── Frontend ────────────────────────────────────

@app.route("/")
def index():
    resp = send_from_directory(STATIC_DIR, "index.html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


def _zpl_inventario(receiptkey: str, arm_display: str, nivel_str: str,
                    dpi: int = 203, width_mm: float = 95.0, height_mm: float = 90.0) -> bytes:
    """Gera ZPL para etiqueta de inventário com dimensões exatas 95×90mm."""
    def dots(mm: float) -> int:
        return max(1, round(mm * dpi / 25.4))

    W = dots(width_mm)
    H = dots(height_mm)
    pad = dots(3)

    f_logo = dots(6.0)
    f_arm  = dots(7.5)
    f_cont = dots(4.5)
    f_ban  = dots(3.5)
    f_doc  = dots(3.5)
    f_dt   = dots(2.5)

    from datetime import datetime as _dt
    agora = _dt.now().strftime("%d/%m/%Y %H:%M")

    ln = ["^XA", f"^PW{W}", f"^LL{H}", "^CI28", "^LH0,0", "^MD5"]
    y = pad

    ln += [f"^FO0,{y}^A0N,{f_logo},{f_logo}^FB{W},1,0,C^FDLIMPPANO^FS"]
    y += f_logo + dots(2)

    ln += [f"^FO0,{y}^GB{W},3,3^FS"]
    y += dots(2)

    ln += [f"^FO0,{y}^A0N,{f_arm},{f_arm}^FB{W},1,0,C^FD{arm_display}^FS"]
    y += f_arm + dots(1)

    ln += [f"^FO0,{y}^A0N,{f_cont},{f_cont}^FB{W},1,0,C^FDCONTAGEM {nivel_str}^FS"]
    y += f_cont + dots(1.5)

    ln += [f"^FO0,{y}^GB{W},3,3^FS"]
    y += dots(2)

    ban_h = f_ban + dots(6)
    ln += [
        f"^FO0,{y}^GB{W},{ban_h},{ban_h}^FS",
        f"^FO0,{y + (ban_h - f_ban) // 2}^A0N,{f_ban},{f_ban}^FB{W},1,0,C^FR^FDINVENTARIO^FS",
    ]
    y += ban_h + dots(2)

    text_area = f_doc + dots(2) + f_dt + dots(2)
    qr_area = max(dots(15), H - pad - y - text_area)
    modules = 33  # QR type 4 (≤34 chars, M correction) = 33×33 módulos
    mag = max(2, min(10, qr_area // modules))
    qr_px = modules * mag
    qr_x = max(0, (W - qr_px) // 2)
    qr_y = y + max(0, (qr_area - qr_px) // 2)
    ln += [f"^FO{qr_x},{qr_y}^BQN,2,{mag}^FDQA,{receiptkey}^FS"]
    y += qr_area

    ln += [f"^FO0,{y}^A0N,{f_doc},{f_doc}^FB{W},1,0,C^FD{receiptkey}^FS"]
    y += f_doc + dots(1.5)
    ln += [f"^FO0,{y}^A0N,{f_dt},{f_dt}^FB{W},1,0,C^FD{agora}^FS"]
    ln += ["", "^PQ1", "^XZ"]
    return "\n".join(ln).encode("utf-8")


@app.get("/api/impressoras")
def api_impressoras():
    """Lista impressoras Zebra disponíveis (Windows) ou vazio (Linux/Render)."""
    try:
        import win32print
        todas = [info[2] for info in win32print.EnumPrinters(
            win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)]
        zebra = [p for p in todas if any(
            k in p.upper() for k in ("ZEBRA", "ZDESIGNER", "ZD", "ZT", "GK", "GX", "ZPL"))]
        return jsonify(zebra if zebra else todas)
    except Exception:
        return jsonify([])


@app.route("/etiqueta/<receiptkey>/imprimir", methods=["POST"])
def etiqueta_imprimir(receiptkey: str):
    """Envia ZPL direto para a Zebra (Windows local) com dimensões exatas 95×90mm."""
    try:
        import win32print
    except ImportError:
        return jsonify({"ok": False, "erro": "Impressão direta indisponível neste servidor."}), 400
    try:
        data = request.get_json(silent=True) or {}
        printer_name = data.get("impressora", "")
        if not printer_name:
            return jsonify({"ok": False, "erro": "Impressora não especificada."}), 400

        parts = receiptkey.split("_CONTAGEM")
        nivel_str = parts[1] if len(parts) > 1 else "?"
        arm = parts[0].split("_")[-1] if parts else "?"
        arm_display = (arm[:4] + " - " + arm[4:]) if len(arm) > 4 else arm

        zpl = _zpl_inventario(receiptkey, arm_display, nivel_str)

        ph = win32print.OpenPrinter(printer_name)
        try:
            win32print.StartDocPrinter(ph, 1, ("Etiqueta Inventario", None, "RAW"))
            win32print.StartPagePrinter(ph)
            win32print.WritePrinter(ph, zpl)
            win32print.EndPagePrinter(ph)
            win32print.EndDocPrinter(ph)
        finally:
            win32print.ClosePrinter(ph)

        log.info(f"ZPL enviado para '{printer_name}': {receiptkey}")
        return jsonify({"ok": True})
    except Exception as e:
        log.error(f"Erro ao imprimir ZPL: {e}")
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.get("/etiqueta/<receiptkey>/zpl")
def etiqueta_zpl(receiptkey: str):
    """Retorna ZPL raw para uso via Zebra Browser Print (qualquer origem)."""
    parts = receiptkey.split("_CONTAGEM")
    nivel_str = parts[1] if len(parts) > 1 else "?"
    arm = parts[0].split("_")[-1] if parts else "?"
    arm_display = (arm[:4] + " - " + arm[4:]) if len(arm) > 4 else arm
    zpl = _zpl_inventario(receiptkey, arm_display, nivel_str)
    from flask import Response
    resp = Response(zpl, mimetype="text/plain")
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp


@app.route("/etiqueta/<receiptkey>")
def etiqueta(receiptkey: str):
    """Gera etiqueta 95×90mm com QR code embutido e impressão direta ZPL."""
    import io, base64
    import qrcode as _qr
    from markupsafe import escape
    rk = str(escape(receiptkey))
    parts = rk.split("_CONTAGEM")
    nivel_str = parts[1] if len(parts) > 1 else "?"
    arm = parts[0].split("_")[-1] if parts else "?"
    arm_display = (arm[:4] + " - " + arm[4:]) if len(arm) > 4 else arm
    from datetime import datetime as _dt
    agora = _dt.now().strftime("%d/%m/%Y %H:%M")

    qr = _qr.QRCode(error_correction=_qr.constants.ERROR_CORRECT_M, box_size=6, border=2)
    qr.add_data(receiptkey)
    qr.make(fit=True)
    buf = io.BytesIO()
    qr.make_image(fill_color="black", back_color="white").save(buf, format="PNG")
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    rk_js = receiptkey.replace("\\", "\\\\").replace('"', '\\"')

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Etiqueta — {rk}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#e8eef7;
        display:flex;flex-direction:column;align-items:center;padding:16px}}
  .controls{{background:#1A5276;color:#fff;padding:8px 14px;border-radius:8px;
    margin-bottom:14px;display:flex;align-items:center;gap:8px;width:95mm;flex-wrap:wrap}}
  .ctrl-info{{flex:1;font-size:10px;min-width:80px}}
  .sel-imp{{background:#fff;color:#1A5276;border:none;padding:4px 6px;border-radius:5px;
    font-size:11px;font-weight:700;cursor:pointer;display:none;max-width:130px}}
  .btn-p{{background:#fff;color:#1A5276;border:none;padding:5px 12px;border-radius:6px;
    font-weight:700;font-size:12px;cursor:pointer;white-space:nowrap}}
  .btn-p:hover{{background:#d6eaf8}}
  .btn-p:disabled{{opacity:.6;cursor:wait}}
  .label{{width:95mm;height:90mm;background:#fff;border:1.5px solid #17202a;
    display:flex;flex-direction:column;padding:3mm;overflow:hidden}}
  .logo-txt{{text-align:center;font-size:6mm;font-weight:900;color:#1A5276;
    letter-spacing:2px;line-height:1;padding-bottom:1.5mm;flex-shrink:0}}
  .sep{{height:0.5mm;background:#17202a;flex-shrink:0}}
  .arm{{text-align:center;font-size:7.5mm;font-weight:900;color:#17202a;
    letter-spacing:2px;padding:1mm 0 0;line-height:1.1;flex-shrink:0}}
  .contagem{{text-align:center;font-size:4.5mm;font-weight:700;color:#17202a;
    letter-spacing:2px;padding:0.5mm 0 1mm;line-height:1;flex-shrink:0}}
  .banda{{background:#17202a;color:#fff;text-align:center;font-size:3.5mm;
    font-weight:700;letter-spacing:2px;padding:1.5mm 0;margin:1mm 0;flex-shrink:0}}
  .qr-wrap{{flex:1;min-height:0;display:flex;justify-content:center;
    align-items:center;overflow:hidden}}
  .qr-wrap img{{display:block;max-height:40mm;max-width:40mm;width:auto;height:auto;
    image-rendering:pixelated}}
  .doc{{text-align:center;font-size:3.5mm;font-weight:700;color:#17202a;
    font-family:'Courier New',monospace;letter-spacing:0.3mm;padding:1mm 0 0;
    word-break:break-all;line-height:1.2;flex-shrink:0}}
  .dt{{text-align:center;font-size:2.2mm;color:#5D6D7E;padding-top:0.8mm;flex-shrink:0}}
  @page{{size:95mm 90mm;margin:0}}
  @media print{{body{{background:#fff;padding:0}}.controls{{display:none}}
    .label{{border:none;width:95mm;height:90mm}}}}
</style>
</head>
<body>
  <div class="controls">
    <span class="ctrl-info">C{nivel_str} — {arm_display}&nbsp;|&nbsp;{agora}</span>
    <select id="sel-imp" class="sel-imp"></select>
    <button class="btn-p" id="btn-imp">&#128424; Imprimir</button>
  </div>
  <div class="label">
    <div class="logo-txt">LIMPPANO</div>
    <div class="sep"></div>
    <div class="arm">{arm_display}</div>
    <div class="contagem">CONTAGEM {nivel_str}</div>
    <div class="sep"></div>
    <div class="banda">INVENTÁRIO</div>
    <div class="qr-wrap">
      <img src="data:image/png;base64,{qr_b64}" alt="QR Code">
    </div>
    <div class="doc">{rk}</div>
    <div class="dt">{agora}</div>
  </div>
<script>
const RK = "{rk_js}";
const SEL = document.getElementById('sel-imp');
const BTN = document.getElementById('btn-imp');
const ZBP  = 'http://localhost:9100';

async function carregarImpressoras() {{
  try {{
    const r = await fetch('/api/impressoras');
    const lst = await r.json();
    if (lst && lst.length > 0) {{
      SEL.style.display = 'block';
      lst.forEach(function(p) {{
        const o = document.createElement('option');
        o.value = p; o.textContent = p.length > 22 ? p.substring(0,20)+'…' : p;
        o.title = p;
        SEL.appendChild(o);
      }});
    }}
  }} catch(e) {{}}
}}

async function zbpImprimir() {{
  // Busca ZPL do servidor Render/local
  const zplR = await fetch('/etiqueta/' + encodeURIComponent(RK) + '/zpl');
  const zpl  = await zplR.text();
  // Conecta ao Zebra Browser Print rodando localmente
  const ctrl  = new AbortController();
  const timer = setTimeout(function() {{ ctrl.abort(); }}, 3000);
  let pbData;
  try {{ pbData = await (await fetch(ZBP + '/available', {{signal: ctrl.signal}})).json(); }}
  finally {{ clearTimeout(timer); }}
  const imp = pbData && pbData.printer && pbData.printer[0];
  if (!imp) throw new Error('ZBP: nenhuma impressora');
  const wr = await fetch(ZBP + '/write', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{device: imp, data: zpl}})
  }});
  if (!wr.ok) throw new Error('ZBP write: ' + wr.status);
  return imp.name || 'Zebra';
}}

function resetBtn() {{
  BTN.disabled = false; BTN.textContent = '🖨 Imprimir';
  BTN.style.background = ''; BTN.style.color = '';
}}

async function doImprimir() {{
  BTN.disabled = true;

  // Prioridade 1: impressora local via servidor Flask
  if (SEL.value) {{
    BTN.textContent = '⏳ Enviando...';
    try {{
      const r = await fetch('/etiqueta/' + encodeURIComponent(RK) + '/imprimir', {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify({{impressora: SEL.value}})
      }});
      const d = await r.json();
      if (d.ok) {{
        BTN.textContent = '✓ Impresso!';
        BTN.style.background = '#a9dfbf'; BTN.style.color = '#1e8449';
      }} else {{
        alert('Erro: ' + d.erro); resetBtn(); return;
      }}
    }} catch(e) {{ alert('Falha: ' + e); resetBtn(); return; }}
    setTimeout(resetBtn, 3000);
    return;
  }}

  // Prioridade 2: Zebra Browser Print (funciona do Render via localhost)
  BTN.textContent = '⏳ ZBP...';
  try {{
    await zbpImprimir();
    BTN.textContent = '✓ Impresso!';
    BTN.style.background = '#a9dfbf'; BTN.style.color = '#1e8449';
    setTimeout(resetBtn, 3000);
    return;
  }} catch(e) {{
    // ZBP não disponível — fallback para janela de impressão do browser
  }}

  resetBtn();
  window.print();
}}

BTN.addEventListener('click', doImprimir);
carregarImpressoras();
</script>
</body>
</html>"""
    from flask import Response
    return Response(html, mimetype="text/html")


@app.get("/swagger.json")
def swagger_json():
    """OpenAPI/Swagger definition — usado pelo ION API Gateway para registrar operações."""
    spec = {
        "swagger": "2.0",
        "info": {
            "title": "PainelRecebimentoPRD",
            "description": "Painel de Recebimento de Produto Acabado BURN RJ",
            "version": "1.0.0",
        },
        "host": "painel-burn-rj.onrender.com",
        "basePath": "/api",
        "schemes": ["https"],
        "paths": {
            "/webhook-asn": {
                "post": {
                    "summary": "Notificacao de nova ASN",
                    "description": "Recebe notificacao do ION DataFlow quando uma ASN e criada ou atualizada no WMS",
                    "operationId": "postWebhookAsn",
                    "consumes": ["application/json"],
                    "produces": ["application/json"],
                    "parameters": [
                        {
                            "in": "body",
                            "name": "body",
                            "required": True,
                            "schema": {
                                "type": "object",
                                "properties": {
                                    "receiptkey": {
                                        "type": "string",
                                        "description": "Chave do recebimento no WMS",
                                    }
                                },
                            },
                        }
                    ],
                    "responses": {
                        "200": {"description": "ASN indexada com sucesso"},
                        "400": {"description": "receiptkey ausente no body"},
                        "404": {"description": "ASN nao encontrada no WMS"},
                    },
                }
            }
        },
    }
    return jsonify(spec)


# ─────────────────────────────── API ─────────────────────────────────────────

@app.get("/api/farol")
def api_farol():
    """Retorna o resumo do farol agrupado por depósito e status."""
    data_filtro = request.args.get("data")  # YYYY-MM-DD opcional
    try:
        farol = collector.get_farol(data_filtro=data_filtro)
        return jsonify(farol)
    except Exception as e:
        log.error(f"Erro em /api/farol: {e}")
        return jsonify({"erro": str(e)}), 500


@app.get("/api/receipts")
def api_receipts():
    """
    Lista receipts com filtros opcionais:
      deposito=308&status=pendente&bucket=hoje|backlog|todos
    """
    deposito = request.args.get("deposito", "")
    status   = request.args.get("status", "")
    bucket   = request.args.get("bucket", "todos")
    try:
        items = collector.get_receipts_by_status_deposito(deposito, status, bucket)
        return jsonify({"items": items, "total": len(items)})
    except Exception as e:
        log.error(f"Erro em /api/receipts: {e}")
        return jsonify({"erro": str(e)}), 500


@app.get("/api/receipt/<receiptkey>")
def api_receipt_detail(receiptkey):
    """Retorna detalhe completo de um recebimento."""
    try:
        rec = collector.get_receipt_detail(receiptkey)
        if rec is None:
            return jsonify({"erro": "Receipt não encontrado"}), 404
        return jsonify(rec)
    except Exception as e:
        log.error(f"Erro em /api/receipt/{receiptkey}: {e}")
        return jsonify({"erro": str(e)}), 500


@app.post("/api/refresh")
def api_refresh():
    """Força atualização imediata dos dados."""
    try:
        collector.forcar_refresh()
        return jsonify({"ok": True, "msg": "Refresh iniciado em background."})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.get("/api/fetch-asn/<receiptkey>")
def api_fetch_asn(receiptkey):
    """Busca e indexa uma ASN específica imediatamente pelo receiptkey."""
    try:
        rec = collector.fetch_and_store(receiptkey)
        if rec:
            return jsonify({"ok": True, "receipt": rec})
        return jsonify({"ok": False, "msg": f"ASN {receiptkey} não encontrada no WMS."}), 404
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.get("/api/depositos")
def api_depositos():
    """Retorna lista de depósitos configurados."""
    return jsonify(DEPOSITOS)


@app.get("/api/status")
def api_status():
    """Health check e status do coletor."""
    state = collector.get_state()
    return jsonify({
        "ok":                  state["erro"] is None,
        "ultima_atualizacao":  state["ultima_atualizacao"],
        "total_receipts":      len(state["receipts"]),
        "erro":                state["erro"],
        "server_time":         datetime.now(_BRT).strftime("%d/%m/%Y %H:%M:%S"),
    })


@app.get("/api/stream")
def api_stream():
    """SSE — empurra evento para o browser sempre que o coletor atualiza dados."""
    def _generate():
        last_ts = None
        # keepalive: comenta vazia a cada 15s para não quebrar proxies/Render
        idle = 0
        while True:
            state = collector.get_state()
            ts = state.get("ultima_atualizacao") or ""
            if ts != last_ts:
                last_ts = ts
                idle = 0
                yield f"data: {ts}\n\n"
            else:
                idle += 1
                if idle >= 5:       # 5 × 3s = 15s sem mudança → keepalive
                    idle = 0
                    yield ": keepalive\n\n"
            time.sleep(3)
    return Response(
        _generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/debug-wms-list")
def api_debug_wms_list():
    """Testa métodos de listagem de ASNs — descobre qual endpoint funciona."""
    from receipt_collector import WMSClient
    from datetime import date
    client = WMSClient()
    dt = date.today().isoformat()
    resultado = {}

    # 1. GET /advancedshipnotice sem filtro de data (apenas recordcount)
    for path in ("advancedshipnotice", "receipts"):
        try:
            r = client.get(path, params={"storerkey": "BURN", "recordcount": 10}, timeout=30)
            keys, sample = [], []
            if r.status_code == 200:
                data = r.json()
                if isinstance(data, list):
                    keys = [x.get("receiptkey") for x in data[:5]]
                    sample = [{k: v for k,v in x.items() if "date" in k.lower() or k in ("receiptkey","status","supplierCode","SupplierCode")} for x in data[:2]]
            resultado[f"GET_{path}_nofilter"] = {"status": r.status_code, "keys": keys, "sample": sample, "body": r.text[:300] if r.status_code != 200 else ""}
        except Exception as e:
            resultado[f"GET_{path}_nofilter"] = {"erro": str(e)}

    # 2. GET por status aberto (sem data)
    for status_val in ("5", "9", "3"):
        try:
            r = client.get("advancedshipnotice", params={"storerkey": "BURN", "status": status_val, "recordcount": 10}, timeout=30)
            keys = []
            if r.status_code == 200 and isinstance(r.json(), list):
                keys = [x.get("receiptkey") for x in r.json()[:10]]
            resultado[f"GET_asn_status_{status_val}"] = {"status": r.status_code, "keys": keys, "body": r.text[:200] if r.status_code != 200 else ""}
        except Exception as e:
            resultado[f"GET_asn_status_{status_val}"] = {"erro": str(e)}

    # 3. GET por editdate (data da atualização)
    for path in ("advancedshipnotice", "receipts"):
        for param in ("editdate", "lastmoddate", "updatedate", "modifieddate"):
            try:
                r = client.get(path, params={"storerkey": "BURN", param: dt, "recordcount": 10}, timeout=15)
                keys = []
                if r.status_code == 200 and isinstance(r.json(), list):
                    keys = [x.get("receiptkey") for x in r.json()[:5]]
                if r.status_code != 405:  # só mostra se não for 405
                    resultado[f"GET_{path}_{param}"] = {"status": r.status_code, "keys": keys, "body": r.text[:200]}
            except Exception as e:
                resultado[f"GET_{path}_{param}"] = {"erro": str(e)}

    # 4. Exports — ver último evento e total
    try:
        r = client.get("exports", params={"type": "ASNCOMPLETED", "restrictrowsto": 5}, timeout=30)
        sample = []
        if r.status_code == 200 and isinstance(r.json(), list):
            sample = [{"key1": x.get("key1","")[:30], "adddate": x.get("adddate","")} for x in r.json()]
        resultado["exports_ASNCOMPLETED_5"] = {"status": r.status_code, "sample": sample}
    except Exception as e:
        resultado["exports_ASNCOMPLETED_5"] = {"erro": str(e)}

    # 5. Inventário nos stages (scan direto)
    try:
        r = client.post("inventorybalance/showinventorybalancelist",
                        params={"recordcount": 5, "loc": "STG.PA.309.001", "owner": "BURN"}, timeout=30)
        fields = []
        if r.status_code == 200 and isinstance(r.json(), list) and r.json():
            fields = list(r.json()[0].keys())
            sample = [{k: v for k, v in r.json()[0].items() if v and v != "0" and v != ""}]
        resultado["stage_309_sample"] = {"status": r.status_code, "fields": fields[:30], "sample": sample if r.status_code == 200 else r.text[:200]}
    except Exception as e:
        resultado["stage_309_sample"] = {"erro": str(e)}

    return jsonify(resultado)


@app.get("/api/debug-asn-raw/<receiptkey>")
def api_debug_asn_raw(receiptkey):
    """Retorna todos os campos brutos do cabeçalho da ASN (sem expandir detalhes)."""
    from receipt_collector import WMSClient
    client = WMSClient()
    r = client.get(f"advancedshipnotice/{receiptkey}", timeout=30)
    if r.status_code == 200:
        raw = r.json()
        # Remove campos muito grandes para leitura
        resumo = {k: v for k, v in raw.items() if k not in ("receiptdetails", "fieldOverrides", "hrefs", "facility", "link")}
        return jsonify({"status_wms": 200, "campos": resumo})
    return jsonify({"status_wms": r.status_code, "body": r.text[:500]}), 404


@app.get("/api/debug-pack/<packkey>")
def api_debug_pack(packkey):
    """Retorna todos os campos brutos do cadastro de embalagem (pack) do WMS."""
    from receipt_collector import WMSClient
    client = WMSClient()
    r = client.get(f"packs/{packkey}", timeout=20)
    if r.status_code == 200:
        raw = r.json()
        # Calcular qpp como o sistema faz
        ti = float(raw.get("palletti") or 0)
        hi = float(raw.get("pallethi") or 0)
        qty = float(raw.get("qty") or 0)
        qpp_calc = ti * hi if (ti > 0 and hi > 0) else (qty if qty > 0 else 0)
        return jsonify({
            "packkey":    packkey,
            "status_wms": r.status_code,
            "qpp_calculado": qpp_calc,
            "formula":    f"palletti({ti}) × pallethi({hi}) = {ti*hi}" if (ti > 0 and hi > 0) else f"fallback qty={qty}",
            "campos_raw": raw,
        })
    return jsonify({"status_wms": r.status_code, "body": r.text[:500]}), 404


@app.get("/api/debug-receipt")
def api_debug_receipt():
    """Diagnóstico: campos chave de até 5 receipts — ajuda a identificar deposito/fromloc/toloc."""
    state = collector.get_state()
    receipts = state["receipts"]
    amostra = []
    for rk, rec in list(receipts.items())[:5]:
        linhas = rec.get("linhas") or []
        linhas_info = [
            {
                "fromloc":      l.get("fromloc"),
                "toloc":        l.get("toloc"),
                "packkey":      l.get("packkey"),
                "qty_por_palete": l.get("qty_por_palete"),
                "qty_previsto": l.get("qty_previsto"),
                "qty_recebido": l.get("qty_recebido"),
            }
            for l in linhas[:3]
        ]
        amostra.append({
            "receiptkey":       rk,
            "externkey":        rec.get("externkey"),
            "deposito":         rec.get("deposito"),
            "supplier_code":    rec.get("supplier_code"),
            "status":           rec.get("status"),
            "status_wms":       rec.get("status_wms"),
            "status_raw":       rec.get("status_raw"),
            "data_criacao":     rec.get("data_criacao"),
            "data_recebimento": rec.get("data_recebimento"),
            "data_fechamento":  rec.get("data_fechamento"),
            "n_linhas":         rec.get("n_linhas"),
            "paletes_total":    rec["paletes"]["paletes_total"] if rec.get("paletes") else None,
            "linhas":           linhas_info,
        })
    return jsonify(amostra)


@app.post("/api/bulk-import")
def api_bulk_import():
    """Importa lista de ASNs pré-parseadas diretamente no cache (sem chamar o WMS)."""
    try:
        body = request.get_json(force=True, silent=True) or {}
        receipts = body.get("receipts", [])
        if not receipts:
            return jsonify({"ok": False, "msg": "lista receipts vazia"}), 400
        imported = collector.bulk_import(receipts)
        return jsonify({"ok": True, "imported": imported})
    except Exception as e:
        log.error(f"Erro em /api/bulk-import: {e}")
        return jsonify({"erro": str(e)}), 500


_webhook_last_body: dict = {}   # último payload recebido — para diagnóstico


def _extrair_receiptkey(body: dict) -> str:
    """
    Tenta extrair o receiptkey de múltiplos formatos possíveis de payload:
      1. JSON plano do WMS  { "receiptkey": "..." }
      2. BOD SyncAdvanceShipNotice  DataArea.AdvanceShipNotice[0].AdvanceShipNoticeHeader.ID.value
      3. Campos alternativos de referência externa (externreceiptkey, DocumentID, etc.)
    """
    if not body:
        return ""

    # 1. JSON plano do WMS (chave direta)
    for k in ("receiptkey", "ReceiptKey", "receiptKey", "externreceiptkey", "ExternReceiptKey"):
        v = str(body.get(k) or "").strip()
        if v:
            return v

    # 2. BOD SyncAdvanceShipNotice — DataArea.AdvanceShipNotice[].AdvanceShipNoticeHeader.ID
    try:
        data_area = body.get("DataArea") or {}
        asns = data_area.get("AdvanceShipNotice") or []
        if isinstance(asns, dict):
            asns = [asns]
        for asn in asns:
            hdr = asn.get("AdvanceShipNoticeHeader") or {}
            # ID principal
            v = (hdr.get("ID") or {}).get("value", "")
            if v:
                return str(v).strip()
            # Referência de documento (externreceiptkey)
            doc_ref = hdr.get("DocumentReference") or {}
            v = (doc_ref.get("DocumentID") or {}).get("value", "")
            if v:
                return str(v).strip()
    except Exception:
        pass

    # 3. Qualquer chave que contenha "receiptkey" (case-insensitive)
    for k, v in body.items():
        if "receiptkey" in k.lower() and v:
            return str(v).strip()

    return ""


@app.post("/api/webhook-asn")
def api_webhook_asn():
    """
    Webhook para receber notificações do ION DataFlow SyncAdvanceShipNotice.
    Aceita JSON plano do WMS e formato BOD SyncAdvanceShipNotice.
    Sempre retorna 200 para evitar retries do DataFlow.
    """
    global _webhook_last_body
    try:
        raw = request.get_data(as_text=True)
        body = request.get_json(force=True, silent=True) or {}
        _webhook_last_body = {
            "ts":      datetime.now(_BRT).strftime("%d/%m/%Y %H:%M:%S"),
            "headers": dict(request.headers),
            "body":    body,
            "raw":     raw[:2000],
        }
        log.info(f"webhook-asn recebido: {raw[:500]}")

        receiptkey = _extrair_receiptkey(body)
        if not receiptkey:
            log.warning(f"webhook-asn: não foi possível extrair receiptkey do payload")
            # Retorna 200 mesmo assim para não gerar erro no DataFlow
            return jsonify({"ok": False, "msg": "receiptkey não encontrado no payload", "received_keys": list(body.keys())})

        rec = collector.fetch_and_store(receiptkey)
        if rec:
            log.info(f"webhook-asn: ASN {receiptkey} indexada (dep={rec.get('deposito')}, status={rec.get('status')}).")
            return jsonify({"ok": True, "receiptkey": receiptkey, "deposito": rec.get("deposito"), "status": rec.get("status")})

        log.warning(f"webhook-asn: ASN {receiptkey} não encontrada no WMS")
        return jsonify({"ok": False, "msg": f"ASN {receiptkey} não encontrada no WMS"})

    except Exception as e:
        log.error(f"Erro em /api/webhook-asn: {e}")
        return jsonify({"erro": str(e)})   # 200 mesmo em exceção


@app.get("/api/webhook-debug")
def api_webhook_debug():
    """Retorna o último payload recebido em /api/webhook-asn — para diagnóstico do DataFlow."""
    return jsonify(_webhook_last_body or {"msg": "Nenhum payload recebido ainda"})


# ─────────────────────────────── Inventário ──────────────────────────────────

_inventario_dados = {}  # dados em memória — recarregados via POST /api/inventario/dados
_painel_dados     = {}  # dados do painel ERP×WMS — recarregados via POST /api/inventario/painel
_finalizacoes     = {}  # registros de finalização por nível — POST /api/inventario/finalizar
_pending_abrir      = None   # timestamp do comando "abrir inventário" aguardando execução local
_pending_finalizar  = None   # timestamp do comando "finalizar inventário" aguardando execução local
_inventario_final   = {}     # dados do relatório final (pdf gerado, timestamp)

# Arquivo de persistência de pending (sobrevive a restarts de dyno no Render)
_PENDING_FILE = os.path.join(os.path.dirname(__file__), "pending_inv.json")
# Cache do painel ERP×WMS em disco — sobrevive a restarts do dyno Render
_PAINEL_CACHE_FILE = os.path.join(os.path.dirname(__file__), "painel_dados_cache.json")


def _painel_save():
    """Persiste _painel_dados em arquivo para sobreviver a restarts."""
    try:
        with open(_PAINEL_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(_painel_dados, f, ensure_ascii=False)
        log.info(f"Painel persistido em {_PAINEL_CACHE_FILE}")
    except Exception as e:
        log.warning(f"Não foi possível salvar cache do painel: {e}")


def _painel_load():
    """Restaura _painel_dados do cache em disco após restart."""
    global _painel_dados
    if not os.path.exists(_PAINEL_CACHE_FILE):
        return
    if _painel_dados:  # já carregado via POST, não sobrescreve
        return
    try:
        data = json.loads(open(_PAINEL_CACHE_FILE, encoding="utf-8").read())
        if data:
            _painel_dados = data
            log.info(f"Painel restaurado do cache: {data.get('total_skus',0)} SKUs, "
                     f"gerado_em={data.get('gerado_em','?')}")
    except Exception as e:
        log.warning(f"Não foi possível restaurar cache do painel: {e}")


def _pending_save():
    """Persiste os comandos pendentes em arquivo."""
    try:
        data = {}
        if _pending_abrir:
            data["abrir"] = _pending_abrir
        if _pending_finalizar:
            data["finalizar"] = _pending_finalizar
        if data:
            with open(_PENDING_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f)
        elif os.path.exists(_PENDING_FILE):
            os.remove(_PENDING_FILE)
    except Exception as e:
        log.warning(f"Não foi possível salvar pending: {e}")


def _pending_load():
    """Carrega pending persistido do arquivo (chamado na primeira poll após reinício)."""
    global _pending_abrir, _pending_finalizar
    if not os.path.exists(_PENDING_FILE):
        return
    try:
        data = json.loads(open(_PENDING_FILE, encoding="utf-8").read())
        if data.get("abrir") and not _pending_abrir:
            _pending_abrir = data["abrir"]
            log.info(f"Pending 'abrir' restaurado do arquivo: {_pending_abrir}")
        if data.get("finalizar") and not _pending_finalizar:
            _pending_finalizar = data["finalizar"]
            log.info(f"Pending 'finalizar' restaurado do arquivo: {_pending_finalizar}")
    except Exception as e:
        log.warning(f"Não foi possível carregar pending: {e}")


@app.post("/api/inventario/dados")
def api_inventario_upload():
    """Recebe o inventario_processado.json gerado pelo script local e armazena em memória."""
    try:
        dados = request.get_json(force=True, silent=True) or {}
        if not dados:
            return jsonify({"ok": False, "msg": "body vazio"}), 400
        global _inventario_dados
        _inventario_dados = dados
        _inventario_dados["recebido_em"] = datetime.now(_BRT).isoformat(timespec="seconds")
        log.info(f"Inventário recebido: {dados.get('total_asns',0)} ASNs, {len(dados.get('linhas',[]))} pares loc×sku")
        return jsonify({"ok": True, "total_linhas": len(dados.get("linhas", []))})
    except Exception as e:
        log.error(f"Erro em /api/inventario/dados: {e}")
        return jsonify({"erro": str(e)}), 500


@app.get("/api/inventario")
def api_inventario():
    """Retorna os dados de inventário processados."""
    if not _inventario_dados:
        return jsonify({"vazio": True, "msg": "Nenhum dado de inventário carregado."}), 200
    return jsonify(_inventario_dados)


@app.post("/api/inventario/painel")
def api_painel_upload():
    """Recebe dados do painel ERP×WMS (gerar_dashboard.py) e armazena em memória."""
    try:
        dados = request.get_json(force=True, silent=True) or {}
        if not dados:
            return jsonify({"ok": False, "msg": "body vazio"}), 400
        global _painel_dados
        _painel_dados = dados
        _painel_dados["recebido_em"] = datetime.now(_BRT).isoformat(timespec="seconds")
        log.info(f"Painel recebido: {dados.get('total_skus',0)} SKUs ERP, "
                 f"{dados.get('lidos_end',0)}/{dados.get('total_end',0)} end. lidos, "
                 f"{dados.get('wms_linhas',0)} linhas WMS")
        _painel_save()
        return jsonify({"ok": True, "total_skus": dados.get("total_skus", 0),
                        "total_end": dados.get("total_end", 0)})
    except Exception as e:
        log.error(f"Erro em /api/inventario/painel: {e}")
        return jsonify({"erro": str(e)}), 500


@app.get("/api/inventario/painel")
def api_painel():
    """Retorna os dados do painel ERP×WMS."""
    _painel_load()  # tenta restaurar do cache se ainda vazio
    if not _painel_dados:
        return jsonify({"vazio": True, "msg": "Nenhum dado do painel carregado."}), 200
    resp = dict(_painel_dados)
    resp["finalizacoes"] = _finalizacoes
    return jsonify(resp)


@app.post("/api/inventario/finalizar")
def api_finalizar():
    """Registra a finalização de um nível de contagem."""
    dados = request.get_json(force=True, silent=True) or {}
    nivel = dados.get("nivel")
    armazem = dados.get("armazem", "todos")
    if not nivel:
        return jsonify({"ok": False, "msg": "nivel obrigatorio"}), 400
    global _finalizacoes
    # Chave inclui armazem para não sobrescrever finalizações de outros armazéns no mesmo nível
    chave = f"{nivel}_{armazem}"
    _finalizacoes[chave] = {
        "ts":      datetime.now(_BRT).isoformat(timespec="seconds"),
        "armazem": armazem,
        "nivel":   nivel,
    }
    log.info(f"Contagem C{nivel} finalizada — armazém={armazem} chave={chave}")
    return jsonify({"ok": True, "nivel": nivel})


@app.get("/api/inventario/finalizacoes")
def api_finalizacoes():
    """Retorna todos os registros de finalização."""
    return jsonify(_finalizacoes)


@app.post("/api/inventario/abrir")
def api_abrir_inventario():
    """Sinaliza ao script local para carregar a base ERP e criar as ASNs de C1."""
    global _pending_abrir
    _pending_abrir = datetime.now(_BRT).isoformat(timespec="seconds")
    _pending_save()
    log.info("Comando 'abrir inventário' recebido — aguardando execução local.")
    return jsonify({
        "ok":  True,
        "msg": "Comando recebido. O inventário será aberto na próxima atualização (~1 min).",
        "ts":  _pending_abrir,
    })


@app.post("/api/inventario/cancelar")
def api_cancelar_inventario():
    """Limpa todos os dados de inventário da memória (reset para estado inicial)."""
    global _inventario_dados, _painel_dados, _finalizacoes, _inventario_final
    global _pending_abrir, _pending_finalizar
    _inventario_dados   = {}
    _painel_dados       = {}
    _finalizacoes       = {}
    _inventario_final   = {}
    _pending_abrir      = None
    _pending_finalizar  = None
    _pending_save()
    log.info("Inventário cancelado — todos os dados foram limpos.")
    return jsonify({"ok": True, "msg": "Inventário cancelado. Dados limpos."})


@app.post("/api/inventario/finalizar_total")
def api_finalizar_total():
    """Sinaliza ao script local para gerar o relatório PDF final e enviar por email."""
    global _pending_finalizar
    _pending_finalizar = datetime.now(_BRT).isoformat(timespec="seconds")
    _pending_save()
    log.info("Comando 'finalizar inventário' recebido — aguardando execução local.")
    return jsonify({
        "ok":  True,
        "msg": "Comando recebido. O relatório será gerado e enviado por email em ~1 min.",
        "ts":  _pending_finalizar,
    })


@app.post("/api/inventario/reabrir")
def api_reabrir_inventario():
    """Reabre o inventário finalizado — limpa apenas a finalização, mantém dados de contagem."""
    global _inventario_final, _pending_finalizar
    _inventario_final  = {}
    _pending_finalizar = None
    log.info("Inventário reaberto — finalização removida, dados de contagem mantidos.")
    return jsonify({"ok": True, "msg": "Inventário reaberto. Dados de contagem preservados."})


@app.post("/api/inventario/finalizado")
def api_inventario_finalizado():
    """Recebe confirmação do script local após geração do PDF."""
    global _inventario_final
    dados = request.get_json(force=True, silent=True) or {}
    _inventario_final = {
        "finalizado_em": dados.get("finalizado_em"),
        "pdf":           dados.get("pdf"),
        "ts":            datetime.now(_BRT).isoformat(timespec="seconds"),
    }
    log.info(f"Inventário finalizado — PDF: {dados.get('pdf')}")
    return jsonify({"ok": True})


def _qr_b64(text: str) -> str:
    qr = qrcode.QRCode(version=None, box_size=5, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


@app.get("/inventario/lista/<int:nivel>")
def lista_contagem(nivel):
    """Página imprimível com itens divergentes para recontagem (C2 ou C3)."""
    if nivel not in (2, 3):
        return "Nível inválido (use 2 ou 3)", 400
    if not _painel_dados:
        return "Dados do inventário não disponíveis.", 503

    nivel_anterior = nivel - 1
    prefixo  = _painel_dados.get("prefixo", "INVENTARIO")
    itens    = _painel_dados.get("itens", [])
    linhas   = _painel_dados.get("linhas_wms", [])
    agora    = datetime.now(_BRT).strftime("%d/%m/%Y %H:%M")

    if nivel == 2:
        # C2: divergências normais + Sobras WMS ainda não confirmadas + itens pendentes
        itens_div = [i for i in itens if i.get("status") in {"AGUARDA_C2", "DIVERGENCIA_WMS", "PENDENTE"}]
    else:
        # C3: aguardando C3 + Sobras WMS recontadas em C2 mas ainda com divergência entre C1/C2
        itens_div = [i for i in itens if i.get("status") == "AGUARDA_C3"
                     or (i.get("status") == "DIVERGENCIA_WMS" and i.get("c2") is not None)]

    # Monta índice de localização por (sku, dep_erp) para o nivel anterior
    loc_idx: dict[tuple, list[dict]] = {}
    for l in linhas:
        if l.get("nivel") != nivel_anterior:
            continue
        key = (str(l.get("sku", "")), l.get("dep_erp"))
        loc_idx.setdefault(key, []).append(l)

    # Agrupa itens divergentes por armazem_wms
    grupos: dict[str, list[dict]] = {}
    for it in itens_div:
        arm = it.get("armazem_wms") or it.get("deposito_nome") or "ARMZ"
        grupos.setdefault(arm, []).append(it)

    def _fmt(v, dec=5):
        if v is None: return "—"
        return f"{float(v):,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")

    secoes_html = ""
    for arm in sorted(grupos):
        asn_key  = f"{prefixo}_{arm}_CONTAGEM{nivel}"
        qr_img   = _qr_b64(asn_key)
        itens_arm = grupos[arm]

        # Expande por localização (um item pode estar em múltiplos locs)
        linhas_rel = []
        for it in itens_arm:
            key  = (it["sku"], it.get("deposito"))
            locs = loc_idx.get(key, [])
            if locs:
                for l in sorted(locs, key=lambda x: str(x.get("loc", ""))):
                    qty_ant = it.get(f"c{nivel_anterior}")
                    linhas_rel.append({
                        "loc":     l.get("loc", "—"),
                        "zona":    l.get("zona", ""),
                        "sku":     it["sku"],
                        "qty_erp": it.get("qty_erp"),
                        "qty_ant": qty_ant,
                        "dif":     round(float(qty_ant or 0) - float(it.get("qty_erp") or 0), 5),
                    })
            else:
                qty_ant = it.get(f"c{nivel_anterior}")
                obs = "Não contado em C1" if it.get("status") == "PENDENTE" else ""
                linhas_rel.append({
                    "loc": "—", "zona": obs,
                    "sku": it["sku"],
                    "qty_erp": it.get("qty_erp"),
                    "qty_ant": qty_ant,
                    "dif":  round(float(qty_ant or 0) - float(it.get("qty_erp") or 0), 5),
                })

        linhas_rel.sort(key=lambda x: str(x["loc"]))

        rows = ""
        for r in linhas_rel:
            rows += f"""<tr>
              <td><input type="checkbox" style="width:18px;height:18px"></td>
              <td style="font-family:monospace;font-weight:700">{r['loc']}</td>
              <td style="color:#64748B;font-size:11px">{r['zona']}</td>
              <td style="font-weight:700">{r['sku']}</td>
              <td style="min-width:90px;border-bottom:1px solid #94A3B8">&nbsp;</td>
            </tr>"""

        secoes_html += f"""
        <div class="secao">
          <div class="secao-hdr">
            <div>
              <div class="secao-titulo">C{nivel} — {arm}</div>
              <div class="secao-asn">{asn_key}</div>
              <div class="secao-info">{len(linhas_rel)} endereço(s) divergente(s)</div>
            </div>
            <img src="data:image/png;base64,{qr_img}" class="qr-img" alt="QR {asn_key}">
          </div>
          <table>
            <thead><tr>
              <th style="width:30px"></th>
              <th>Endereço</th><th>Zona</th><th>SKU</th>
              <th style="min-width:90px">Nova Qtd</th>
            </tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Lista Contagem C{nivel} — {prefixo}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1E293B;background:#fff;padding:16px}}
h1{{font-size:16px;font-weight:800;color:#1E3A5F;margin-bottom:2px}}
.sub{{font-size:11px;color:#64748B;margin-bottom:16px}}
.secao{{border:1px solid #CBD5E1;border-radius:8px;margin-bottom:20px;overflow:hidden;page-break-inside:avoid}}
.secao-hdr{{background:#1E3A5F;color:#fff;padding:12px 16px;display:flex;justify-content:space-between;align-items:center;gap:16px}}
.secao-titulo{{font-size:14px;font-weight:800;letter-spacing:.3px}}
.secao-asn{{font-family:monospace;font-size:11px;opacity:.85;margin-top:2px}}
.secao-info{{font-size:10px;opacity:.7;margin-top:4px}}
.qr-img{{width:80px;height:80px;background:#fff;border-radius:4px;padding:3px}}
table{{width:100%;border-collapse:collapse;font-size:11px}}
th{{background:#F1F5F9;padding:6px 10px;text-align:left;font-weight:700;color:#475569;border-bottom:2px solid #CBD5E1}}
td{{padding:6px 10px;border-bottom:1px solid #E2E8F0;vertical-align:middle}}
tr:hover td{{background:#F8FAFC}}
.rodape{{margin-top:24px;font-size:10px;color:#94A3B8;text-align:center;border-top:1px solid #E2E8F0;padding-top:10px}}
@media print{{
  body{{padding:8px}}
  .no-print{{display:none}}
  .secao{{page-break-inside:avoid}}
}}
</style>
</head>
<body>
<div class="no-print" style="margin-bottom:16px">
  <button onclick="window.print()" style="padding:8px 18px;background:#1E3A5F;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:700">🖨 Imprimir</button>
</div>
<h1>Lista de Recontagem — Contagem C{nivel}</h1>
<div class="sub">Inventário: <strong>{prefixo}</strong> &nbsp;|&nbsp; Gerado em: {agora} &nbsp;|&nbsp; {len(itens_div)} SKU(s) divergente(s)</div>
{secoes_html if secoes_html else '<p style="color:#64748B;padding:20px">Nenhum item aguardando C' + str(nivel) + '.</p>'}
<div class="rodape">BURN RJ — Painel de Inventário — Limppano</div>
</body></html>"""

    return Response(html, mimetype="text/html; charset=utf-8")


@app.get("/inventario/enderecos/excel")
def inventario_enderecos_excel():
    """Gera Excel com todos os endereços e status lido/não lido por armazém."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    if not _painel_dados:
        return "Dados não disponíveis. Execute enviar_painel.py primeiro.", 503

    enderecos = _painel_dados.get("enderecos_status", [])
    if not enderecos:
        return "Lista de endereços não encontrada nos dados. Regenere o painel.", 404

    verde  = PatternFill("solid", fgColor="C6EFCE")
    verd_f = Font(color="276221", bold=False)
    verm   = PatternFill("solid", fgColor="FFDFD6")
    verm_f = Font(color="9C0006", bold=False)
    hdr_f  = Font(bold=True, color="FFFFFF")
    hdr_bg = PatternFill("solid", fgColor="1E3A5F")
    borda  = Border(
        bottom=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
    )
    centro = Alignment(horizontal="center")

    wb = Workbook()
    wb.remove(wb.active)

    armazens = sorted(set(e["armazem"] for e in enderecos if e["armazem"]))
    for arm in armazens:
        ws = wb.create_sheet(arm)
        items = [e for e in enderecos if e["armazem"] == arm]
        lidos    = sum(1 for e in items if e["lido"])
        nao_lidos = len(items) - lidos
        pct = round(lidos / len(items) * 100, 1) if items else 0.0

        # Cabeçalho resumo
        ws.merge_cells("A1:E1")
        ws["A1"] = f"{arm} — {lidos} lidos / {nao_lidos} não lidos / {len(items)} total  ({pct}%)"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].fill = PatternFill("solid", fgColor="EFF6FF")
        ws.row_dimensions[1].height = 22

        # Headers
        headers = ["Endereço", "Zona", "Tipo", "Status", "Progresso"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = hdr_f
            c.fill = hdr_bg
            c.alignment = centro

        # Dados
        for row, e in enumerate(items, 3):
            ws.cell(row=row, column=1, value=e["loc"]).font = Font(name="Courier New", size=9)
            ws.cell(row=row, column=2, value=e["zona"])
            ws.cell(row=row, column=3, value=e["tipo"])
            status_val = "Lido" if e["lido"] else "Não lido"
            c_status = ws.cell(row=row, column=4, value=status_val)
            c_status.alignment = centro
            if e["lido"]:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = verde
                c_status.font = verd_f
            else:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = verm
                c_status.font = verm_f
            ws.cell(row=row, column=5)  # vazio (coluna progresso visual)

        # Larguras aproximadas
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 6

        # Congelar linha de cabeçalho
        ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(_BRT).strftime("%Y%m%d_%H%M")
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=enderecos_{ts}.xlsx"},
    )


@app.get("/api/inventario/pending")
def api_inventario_pending():
    """Retorna e limpa comandos pendentes para execução local (chamado pelo auto-update)."""
    global _pending_abrir, _pending_finalizar
    _pending_load()  # restaura do arquivo se reiniciou
    resp = {"abrir_inventario": False, "finalizar_inventario": False}
    if _pending_abrir:
        resp["abrir_inventario"] = True
        resp["ts_abrir"] = _pending_abrir
        _pending_abrir = None
    if _pending_finalizar:
        resp["finalizar_inventario"] = True
        resp["ts_finalizar"] = _pending_finalizar
        _pending_finalizar = None
    _pending_save()  # limpa arquivo após consumir
    return jsonify(resp)


@app.get("/api/config-check")
def api_config_check():
    """Diagnóstico de configuração — mostra prefixo das credenciais (não expõe valores completos)."""
    from wms_config import CONFIG as C
    def mask(v):
        if not v:
            return "(vazio)"
        return v[:12] + "..." + f" [{len(v)} chars]"
    return jsonify({
        "tenant":    C.get("tenant", ""),
        "token_url": C.get("token_url", ""),
        "warehouse": C.get("warehouse", ""),
        "owner":     C.get("owner", ""),
        "ci":        mask(C.get("ci", "")),
        "cs":        mask(C.get("cs", "")),
        "saak":      mask(C.get("saak", "")),
        "sask":      mask(C.get("sask", "")),
    })


# ─────────────────────────────── Startup ─────────────────────────────────────

if __name__ == "__main__":
    log.info("=" * 60)
    log.info("  Gerenciamento de Rotina Logística BURN RJ")
    log.info("  Painel: http://localhost:5002/")
    log.info("  API:    http://localhost:5002/api/farol")
    log.info(f"  Polling a cada {POLL_INTERVALO}s")
    log.info("=" * 60)

    app.run(host="0.0.0.0", port=5002, debug=False, threaded=True)
